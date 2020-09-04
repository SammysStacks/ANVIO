"""
Must Have ANVIO Installed:
	Must be in ANVIO's Enviroment. If on Terminal use: $ conda activate anvio-6.2

Other Packages You Need To Install First:
	$ anvi-setup-ncbi-cogs     # For Genome Annotation in Make Storage
	$ conda install biopython  # For Editing Fasta Files to Remove Non-ATGCN Character in Data Preparation
	$ conda install openpyxl   # For Working with Excel Spreedsheets in Identify Genes
	

The following code is based on the ANVIO outlines:
	http://merenlab.org/2016/11/08/pangenomics-v2/#generating-an-anvio-genomes-storage
"""

# --------------------------------------IMPORTS--------------------------------------------------------------------------#


# Modules Always Needed
import os
# Modules for ANVIO Storage
import multiprocessing
import csv
# To Edit FASTA Files and Remove Extended Nucleotides
from Bio.SeqRecord import SeqRecord
from Bio import Alphabet
from Bio.Seq import Seq
from Bio import SeqIO
from Bio.Alphabet import IUPAC
# Get Imports from Command Line
import sys
# Modules to Format Data in Excel for Blastp Gene Identification
import openpyxl
import openpyxl as xl; 
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
# Create Gene Database for Blastp Gene Identification
import json


# ----------------------------------------------------------------------------------------------------------------#

class helper_Files:
	def __init__(self, output_File_Directory):
		# Make Folder for Pangenome Files if not present. If it already exists, this wont do anything. Folders MUST End with '/'
		self.contigs_Database = output_File_Directory + "contigs_Database/"
		self.fasta_Files =  output_File_Directory + "fasta_Files/"
		self.helper_Folder = output_File_Directory + "Helpers/"
		self.pan_Files =  output_File_Directory + "pan_Files/"
		self.pan_Summary = output_File_Directory + "genome_Data/"
		self.external_genomes = self.contigs_Database + "external-genomes.txt"
		self.genome_storage_database = self.contigs_Database + "storage_Database-GENOMES.db"  # must end with '.db'

	def make_Output_Folders(self, output_File_Directory):
		# Make Folder if not present. If it already exists, this wont do anything
		self.execute_Command("mkdir " + output_File_Directory)
		self.execute_Command("mkdir " + self.contigs_Database)
		self.execute_Command("mkdir " + self.helper_Folder)
		self.execute_Command("mkdir " + self.fasta_Files)
		self.execute_Command("mkdir " + self.pan_Summary)
		self.execute_Command("mkdir " + self.pan_Files)

	def execute_Command(self, command):
		print("\n Executing the Following Command: ", command)
		os.system(command)


class user_Inputs:
	def gather_Inputs(self):
		# Record Variables if Inputed Through Command Line. Example: 'python Run_Full_Pangenome.py Hello Sam' Translates to sys.argv = ['Run_Full_Pangenome.py', 'Hello', 'Sam']
		# Base Case: No Arguments
		if len(sys.argv) == 1:
			return "", "", "", "", ""  # Return Blanks (Using Names in 'if __name__ == "__main__"' at bottom
		# User Inputed Arguments
		print("Program Will Assume you Input Information as:")
		print("python Run_Full_Pangenome.py -in [Input File Directory] -out [Output File Directory] -name [Project Name] -title [Pangenome Title] -T [Threads]\n")
		self.check_Argument(sys.argv[1:], "correct number")
		# Non-Mandatory Inputs in Default State:
		output_File_Directory = "./Output_Files/"
		project_Name = "pangenome_Project"
		pan_Title = "'Pangenome Analysis'"
		threads = "1"
		# Looping through arguments given, taking every other argument after Run_Full_Pangenome.py
		print("We Processed the Following Inputs For You: ", sys.argv[1:])
		for index, arg in enumerate(sys.argv[1::2]):
			# Get Argument
			arg_index = (index+1)*2
			user_Input = sys.argv[arg_index]
			self.check_Argument(user_Input, "general")
			# Match Argument to Variable
			if arg == "-in":
				input_Fasta_File_Directory = self.check_Argument(user_Input, "File")
			elif arg == "-out":
				output_File_Directory = self.check_Argument(user_Input, "File")
			elif arg == "-name":
				project_Name = self.check_Argument(user_Input, "filename")
			elif arg == "-title":
				pan_Title = self.check_Argument(user_Input, "Title")
			elif arg == "-T":
				threads = self.check_Argument(user_Input, "int")
			else:
				print("Incorrect Input Prefix. Please Reformat to the Style Above")
				exit()
		# Record the User Inputs and Excecute Code if Sufficient Inputs Given
		try:
			return input_Fasta_File_Directory, output_File_Directory, project_Name, pan_Title, threads
		# Raise Error if One or More of These Fine Variables is Not Defined
		except:
			print("Error: Missing Input File (-in) Flag, which is Required when Running the Program via the Command line")
			print("Mandatory Minimum Expected: $ python Run_Full_Pangenome.py -in [Input File Directory]")
			exit()
	
	def check_Argument(self, user_Input, stage):
		# Should have an Even Amount of Inputs After the .py File
		if stage == "correct number":
			if len(user_Input)%2 != 0:
				print("You have an ODD Number of Inputs. Did you Add a Flag Without an Input??")
				exit()
		# In General, we do NOT want inputs that start with dashes
		elif stage == "general":
			if user_Input.startswith("-"):
				print("Your Input ", user_Input," Started with a '-', which is NOT Allowed. Please Fix This")
				print("Possible User Error: Did You Place 2 Flags Next to Each Other WITHOUT Their Arguments")
				exit()
		# All Files must end with a Slash
		elif stage == "File":
			if user_Input.endswith("/"):
				return user_Input
			else:
				return user_Input + "/"
		# There Should NOT be Spaces in a Filename for the love of coding (Also no bad characters)
		elif stage == "filename":
			for bad_Char in [" ", ";", ")", "(", "-", ".", "?", "/"]:
				user_Input.replace(bad_Char, "_") # Replace with Underscore
			return user_Input.replace(" ", "_") 
		# If There Are Spaces in the Title, It MUST be Surrounded by Quotes
		elif stage == "Title":
			if " " in user_Input:
				if not user_Input.endswith("'"):
					user_Input = user_Input + "'"
				if not user_Input.startswith("'"):
					user_Input = "'" + user_Input
			return user_Input
		# Threads must be an INTEGER
		elif stage == "int":
			if user_Input.isnumeric():
				return user_Input
			else:
				print("You Inputed a Thread That is NOT AN INTEGER. Lol what does that even mean")
				exit()
		else:
			print("What Argument am I checking?")
			exit()


# ----------------------------------------------------------------------------------------------------------------#

class data_Preparation(helper_Files):
	def __init__(self, input_fasta_file_directory, output_fasta_file_directory):
		super().__init__(output_fasta_file_directory)  # Get Variables Inherited from the helper_Files Class
		self.input_fasta_file_directory = input_fasta_file_directory
		self.dont_touch = set() # If you are nervous, add files you DEFINATELY dont want changed in the variable 'dont_touch'

	def remove_Extended_Nucleotides(self):
		"""
		Removing Non-ATCGN Nucleotides. It should be DNA, so not doing anything with U (unsure if RNA works)
		"""
		print("\nRemoving Non-ATCGN Nucleotides")
		# loop through the files in the directory
		for fasta_file in os.listdir(self.input_fasta_file_directory):
            # Make sure I am only changing FASTA files
			possible_Fasta_Files = tuple([".fasta",  ".fna", ".fsa", ".mpfa", ".fa"])
			if fasta_file not in self.dont_touch and fasta_file.endswith(possible_Fasta_Files):
				# Get the previous name of the File + data
				old_file = self.input_fasta_file_directory + fasta_file
				fasta_sequences = SeqIO.parse(open(old_file), 'fasta')
				# Create new file and write to it
				base = os.path.splitext(fasta_file)[0]
				new_file = self.fasta_Files + base + ".fasta"
				
				records = []
				with open(new_file, "w+") as out_file:
					# Modify the contigs one by one (Using enumerate as online it said it needed to be a generator ??)
					for record in fasta_sequences:
						#print("old \n", record)
						# Get Sequence of the Contig
						new_sequence = str(record.seq)
						# Removing Extented Nucleotides Found via https://www.ncbi.nlm.nih.gov/pmc/articles/PMC2865858/
						for NT in ["R", "Y", "M", "K", "S", "W", "H", "B", "V", "D"]:
							new_sequence = new_sequence.replace(NT,"N")
						record.seq = Seq(new_sequence, record.seq.alphabet)
						#print("new \n", old_record) 
						records.append(record)
						# Write the File
					SeqIO.write(records, new_file, "fasta")
			else:
				print("Not Using: ", fasta_file)

			

	# Fix File Names
	def fix_file_names(self):
		"""
		ANVIO Databases only uses files with a '.fa' extension, NO spaces/non-ASCII characters, and where the first digit is NOT a number
		The following function will check for this and make appropriate changes to the FASTA file names
		"""
		print("\nFixing Names of Files")
		# loop through the files in the directory
		for fasta_file in list(os.listdir(self.fasta_Files)):
			# Make sure I am only changing FASTA files
			fixable_ext = tuple([".fasta",  ".fna", ".fsa", ".mpfa", ".fa"])
			if fasta_file not in self.dont_touch and fasta_file.endswith(fixable_ext):
				# Get the previous name of the File
				old_file = self.fasta_Files + fasta_file
				base = os.path.splitext(fasta_file)[0]
				# Please limit the characters to ASCII letters, digits (not in first position), and the underscore ('_') character
				# This ALSO means no spaces
				for bad_Char in ["'"]:
					base = base.replace(bad_Char, "")  # Delete Character
				for bad_Char in [" ", ";", ")", "(", "-", ".", "?", "/"]:
					base = base.replace(bad_Char, "_") # Replace with Underscore
				# If First letter is a digit, add string
				if base[0].isdigit():
					base = "ANVIO_" + base
				# Change File Extension, must be .fa for ANVIO
				new_file = self.fasta_Files + base + '.fa'
				print(old_file + "   ->   " + new_file)
				os.rename(old_file, new_file)
			else:
				print("Not Using: ", fasta_file)


# ----------------------------------------------------------------------------------------------------------------#

class ANVIO_Storage(helper_Files):
	# Initialize Files
	def __init__(self, output_File_Directory, project_name, threads):
		super().__init__(output_File_Directory)  # Get Variables Inherited from the helper_Files Class
		self.dont_touch = {}  # If you are nervous, add files you DEFINATELY dont want changed in the variable 'dont_touch'
		self.project_name = project_name
		self.threads = int(threads)

	def storage_Step(self, stage):
		# Divide up the work with threading. Threads per loop will be optimally decided to maximize parrell processes
		max_Useful_Pool = sum([Fasta_File.endswith(".fa") for Fasta_File in os.listdir(self.fasta_Files)]) # If the user has more threads than files, its not worth wasting them
		threads_per_loop = min(16, max(self.threads//max_Useful_Pool, 1))  # Not worth going over 16 as the extra threads are not useful
		best_Pool = min(max_Useful_Pool, self.threads)

		pool = multiprocessing.Pool(best_Pool)  # The number of parrelel loops. It helps to maximize this. Note: this will make a lot of intermediate storage
		# Looping through Directory to get Fasta Files (making list() to ensure we use only files PREVIOUSLY present)
		for filename in list(os.listdir(self.fasta_Files)):
			# Make sure I am only using the correct files (NOTE: '-clean.fa' will appear and disappear as the program goes)
			if filename not in self.dont_touch and filename.endswith(".fa") and not filename.endswith("-clean.fa"):
				# Threading means that all loops will be executed at once (cannot rely on each other)
				if stage == "format":
					pool.apply_async(self.format_Fasta_Files, args=(filename,))
				elif stage == "convert":
					pool.apply_async(self.create_Contig_Database, args=(filename, str(threads_per_loop)))
				else:
					print("This Stage Does Not Exist")
					exit()
			else:
				print("Not ", stage, "ing: ", filename)
		pool.close()
		pool.join()

	def format_Fasta_Files(self, fasta_Filename):
		"""
		The following function will remove contigs that have big gaps or small length. It can also simplify the contig names.
		"""
		print("\nReformating Fasta")
		# NOTE: After 2 hours of crying, I finally realized that you will get a 'Not a Fasta' error if you try to overwrite the file (input_File != outfile)
			# The error: anvio.fastalib.FastaLibError: Fasta Lib Error: File './Output_Files/JC_0293_contigs.fa' does not seem to be a FASTA file.
		# Add the '--min-len' flag if you want to exlude contigs below a certain length. Default is 0
		# Add the '--max-percentage-gaps' flag if you want to exlude contigs that have a gap above a certain percent. Default is 100
		# Add the '--simplify-names' flag if you want to simplify contig/scaffold names ('scaffold 1' -> '001'). It helps as the BAM converter will do this with bad deflines arbitrarily
		# Add the '--report-file fasta_deflines_map.txt' flag if you used '--simplify-names' and want to map what was changed (basically use it if you add '--simplify-names')
		min_length = "0"  # Default is 0
		max_gaps = "100"     # Default is 100 (%)
		base = os.path.splitext(fasta_Filename)[0]
		defline_map = self.contigs_Database + base + ".txt"
		input_File = self.fasta_Files + fasta_Filename
		output_File = self.fasta_Files + "hold_" + fasta_Filename
		command = "anvi-script-reformat-fasta '" + input_File + "' --min-len " + min_length + " --max-percentage-gaps " + max_gaps + " -o " + output_File + " --simplify-names --report-file " + defline_map
		self.execute_Command(command)
		# Only keep one of the files
		command = "mv " + output_File + " " + input_File
		self.execute_Command(command)

	def create_Contig_Database(self, filename, threads_per_loop):
		print("\nConverting FASTA to ANVIO Database")
        # Convert FASTA format to ANVIO Database format
		# Add the '--skip-mindful-splitting' flag to not search for optimal splitting and just do the default (makes it run faster)
		# Add the '--split-length' flag to specify which legnth to split (ANVIO will still look for best if '--skip-mindful-splitting' flag is not present. Default is 20,000
		base = os.path.splitext(filename)[0]
		database = base + ".db"
		filename = self.fasta_Files + filename
		command = "anvi-gen-contigs-database -f " + filename + " -o " + self.contigs_Database + database + " -n " + self.project_name
		self.execute_Command(command)
		# Run Hmms
		command = "anvi-run-hmms -c " + self.contigs_Database + database + " --num-threads " + threads_per_loop
		self.execute_Command(command)
        # Annotate the genome: optional
		command = "anvi-run-ncbi-cogs -c " + self.contigs_Database + database + " --num-threads " + threads_per_loop  # This requires you to have previously installed 'anvi-setup-ncbi-cogs'
		self.execute_Command(command)
	
	def create_ANVIO_storage(self):
		# Outfile now has same ending as genomes, so make sure not to place it in itself
		print("\nCreating ANVIO Storage")
		dont_use = {self.genome_storage_database}
		with open(self.external_genomes, 'w+') as storage_file:
			# Make file tab delimited, include ANVIO's header
			writer = csv.writer(storage_file, delimiter='\t')
			writer.writerow(["name", "contigs_db_path"])
			
			# loop through the files in the directory
			for filename in os.listdir(self.contigs_Database):
				#  Make sure I am only using the correct files
				if filename not in dont_use and filename.endswith(".db"):
					# On each row, write the name of file and path to file
					base_name = os.path.splitext(filename)[0]
					ext = os.path.splitext(filename)[-1]
					print(base_name, filename)
					path = "../../" + self.contigs_Database + base_name + ext
					# Write the File. NOTE: The 'path' STARTS (!) from wherever the 'external_genomes' files is
					writer.writerow([base_name, path])
				else:
					print("Not Using: ", filename)
		
		# Run program to generate genome storage
		# optional: Add ' --debug' at the end to see full printout
		command = "anvi-gen-genomes-storage -e " + self.external_genomes + " -o " + self.genome_storage_database
		self.execute_Command(command)
		

# ----------------------------------------------------------------------------------------------------------------#

class Run_Pangenome(helper_Files):
	# Initialize Files
	def __init__(self, output_File_Directory, project_Name, threads):
		super().__init__(output_File_Directory)  # Get Variables Inherited from the helper_Files Class
		self.project_Name = project_Name
		self.threads = str(threads)
	
	def run_Pangenome(self):          
		# Run program to generate pangenome
		# delete '--use-ncbi-blast' if you want to use the faster version of DIAMOND (may be a bit less accurate)
		# Using '--exclude-partial-gene-calls' is optional but removes bad parts in MAG
		# Use '--minbit' to eliminate weak similarities. oprional as the program can recover but applies this constraint at every step
		command = "anvi-pan-genome -g " + self.genome_storage_database + " --project-name " + self.project_Name + " --output-dir " + self.pan_Files + " --num-threads " + self.threads + " --use-ncbi-blast --minbit 0.5 --enforce-hierarchical-clustering"
		self.execute_Command(command)

	def display_Pangenome(self, pan_Title):
		# Run program to generate pangenome
		Pangenome = self.pan_Files + self.project_Name + "-PAN.db"
		command = "anvi-display-pan -g " + self.genome_storage_database + " -p " + Pangenome + " --title " + pan_Title
		self.execute_Command(command)

	def summarize_Pangenome(self, pangenome_Bin_Collection_Saved_As):
		# Summarize the Pangenome you Created
		Pangenome = self.pan_Files + self.project_Name + "-PAN.db"
		pangenome_Summary_Output = self.pan_Summary + pangenome_Bin_Collection_Saved_As + "/" + "Summary/"
		# First List Availible Collections
		command = "anvi-summarize -p " + Pangenome + " -g " + self.genome_storage_database + " --list-collections"
		self.execute_Command(command)
		# Get Collections
		command = "anvi-summarize -p " + Pangenome + " -g " + self.genome_storage_database + " -C " + pangenome_Bin_Collection_Saved_As + " -o " + pangenome_Summary_Output
		self.execute_Command(command)
		# Results May be Zipped
		print("\nUnzipping Summary Files")
		command = "gunzip " + pangenome_Summary_Output + "*"
		self.execute_Command(command)


# ----------------------------------------------------------------------------------------------------------------#
		
class Additions_to_Pangenome(helper_Files):
	"""
	Optional Helper Functions To Add to Pangenome File
	NOTE: You SHOULD DEFINATELY READ what I am doing in each function that you add. Personalize it for YOUR use
	"""
	# Initialize Files
	def __init__(self, output_File_Directory, project_Name, threads):
		super().__init__(output_File_Directory)  # Get Variables Inherited from the helper_Files Class
		self.project_Name = project_Name
		self.threads = str(threads)

	# Add Genome Clusters
	def add_Genome_Clusters(self):
		# If aligning DNA (instead of Amino Acids) add ' --report-DNA-sequence ' at end
		# NOTE: the flag '--concatenate-gene-clusters' may have a bug in it/ I dont understand it
		# Add filters like the following to ensure the clusters are for phylogenomic analysis (no multiple calls): '--max-num-genes-from-each-genome 1'
		pangenome_File = self.pan_Files + self.project_Name + "-PAN.db"
		genome_cluster_output = self.helper_Folder + "genome_clusters.fasta"
		command = "anvi-get-sequences-for-gene-clusters -p " + pangenome_File + " -g " + self.genome_storage_database + " --concatenate-gene-clusters -o " +  genome_cluster_output + " --max-num-genes-from-each-genome 1"
		self.execute_Command(command)

	def add_ANI(self):
		# NOTE: This program will NOT Overwrite Itself (Thankfully!). So you must remove the old ANI before you do a new one
		# Add '--min-full-percent-identity 0.8' if you want to throw out untrustworthy ANI values under 0.8 (see help for more info on this)
		# add '--pan-db YOUR_PANGENOME_DB_FILE' to automatically add these results to pangenome
		min_Identity = "0.75"
		pangenome_File = self.pan_Files + self.project_Name + "-PAN.db"
		command = "anvi-compute-genome-similarity --external-genomes " + self.external_genomes + " --program pyANI --output-dir " + self.helper_Folder + "ANI/" + " --num-threads " + self.threads + " --min-full-percent-identity " + min_Identity + " --pan-db " + pangenome_File
		print("Executing the Following Command: ", command)
		ANVIO_genome_similarity = os.system(command)

	def add_beauty(self):
		# ANVIO provides custom settings to make the pangenome look good
		# NOTE: It may make it look flat, but just redraw as a circle
		import_pretty = "wget http://merenlab.org/tutorials/interactive-interface/files/pretty-state.json"
		self.execute_Command(import_pretty + "; mv pretty-state.json " + self.helper_Folder + "pretty-state.json")
		pangenome_File = self.pan_Files + self.project_Name + "-PAN.db"
		add_lushisness = "anvi-import-state -p " + pangenome_File + " -s " + self.helper_Folder + "pretty-state.json -n default"
		self.execute_Command(add_lushisness)
		

# ----------------------------------------------------------------------------------------------------------------#

class identify_Genes(helper_Files):
	# Initialize Files
	def __init__(self, output_File_Directory, project_Name, pangenome_Bin_Collection_Saved_As, threads):
		super().__init__(output_File_Directory)  # Get Variables Inherited from the helper_Files Class
		self.pangenome_Summary_Output = self.pan_Summary + pangenome_Bin_Collection_Saved_As + "/"
		self.genome_Summary = self.pangenome_Summary_Output + "Summary/" + project_Name + "_gene_clusters_summary.txt"
		self.project_Name = project_Name
		self.threads = int(threads)
		self.dont_touch = {"Summary", "Core", ".DS_Store"}
		
	def compile_Excel(self):
		print("\nCompiling Genes into Excel")
		# Turn txt File into CSV with correct txt delimitter
		genome_Summary_Base = os.path.splitext(self.genome_Summary)[0]
		csv_file = genome_Summary_Base + ".csv"
		# Check to see if csv conversion alreayd happened
		if not os.path.isfile(csv_file):
			with open(self.genome_Summary, "r") as in_text:
				in_reader = csv.reader(in_text, delimiter = '\t')
				with open(csv_file, 'w', newline='') as out_csv:
					out_writer = csv.writer(out_csv)
					for row in in_reader:
						out_writer.writerow(row)
		else:
			print("You already renamed the '.txt' to 'csv'")			
			
		# Get List of Unique Bin Names in Pangenome
		bins = []
		with open(csv_file) as csv_File:
			reader = csv.reader(csv_File, delimiter=',')
			for row in reader:
				# Only Work with Genome Sections that have been Binned
				bin_ID = row[2]
				if bin_ID != "" and bin_ID != "bin_name" and bin_ID not in bins:
					bins.append(bin_ID)

		# Make Excel Workbook for Each Unique Bin Name
		data = {}
		for bin_ID in bins:
			wb = openpyxl.Workbook()
			ws = wb.active
			# Make it an Excel File
			excel_Path = self.pangenome_Summary_Output + bin_ID + "/"
			base_Name = self.genome_Summary.split("/")[-1].split(".")[0]
			self.execute_Command("mkdir " + excel_Path)
			excel_file = excel_Path + base_Name + "_" + bin_ID + ".xlsx"
			# Save to Data
			data[bin_ID] = [wb, ws, excel_file]
			
		# Write to Excel Workbook
		with open(csv_file) as csv_File:
			reader = csv.reader(csv_File, delimiter=',')
			# Loop through CSV and Write to Excel
			for i,row in enumerate(reader):
				bin_ID = row[2]
				# Skip Rows NOT Binned (However, keep the Header/First Row)
				if bin_ID not in bins and i > 0:
					continue
				# Copy Row from CSV
				new_row = []
				for val in row:
					new_row.append(val)
				# Populate to Correct Excel File
				if i == 0:
					for bin_ID in bins:
						ws = data[bin_ID][1]
						ws.append(new_row)  
				else:
					ws = data[bin_ID][1]
					ws.append(new_row)
				
		# Save Excel Workbooks
		for bin_ID in bins:
			wb = data[bin_ID][0]
			excel_file = data[bin_ID][2]
			wb.save(excel_file)
	
	def divide_Work(self, stage, remote = False, species_Name = ""):
		# Divide up the work with threading. Threads per loop will be optimally decided to maximize parrell processes
		max_Useful_Pool = len(os.listdir(self.pangenome_Summary_Output)) - len(self.dont_touch) + 1 # If the user has more threads than files, its not worth wasting them
		threads_per_loop = min(8, max(self.threads//max_Useful_Pool, 1))  # Not worth going over 8 as the extra threads are not useful
		best_Pool = min(max_Useful_Pool, self.threads)
        # If we are using Blastp, we only need 1 thread. max pool should be the maximum (1,000+ proteins)
		if stage == "find_Closest_Protein":
			threads_per_loop = 1
			best_Pool = self.threads

		pool = multiprocessing.Pool(best_Pool)  # The number of parrelel loops. It helps to maximize this. Note: this will make a lot of intermediate storage
		# Looping through Directory to get Bin Folder
		for bin_Folder in list(os.listdir(self.pangenome_Summary_Output)):
			# Make sure I am only using the correct files (NOTE: '-clean.fa' will appear and disappear as the program goes)
			if bin_Folder not in self.dont_touch:
				# Threading means that all loops will be executed at once (cannot rely on each other)
				if stage == "create_Fasta_Files":
					pool.apply_async(self.convert_Gene_Clusters_to_Fasta, args=(bin_Folder,))
				elif stage == "find_Closest_Protein":
					place_Fasta_Files = self.pangenome_Summary_Output + bin_Folder + "/Fasta_Files/"
					self.execute_Command("mkdir " + self.pangenome_Summary_Output + bin_Folder + "/blastp_Results/")
					for fasta_File in os.listdir(place_Fasta_Files):
						pool.apply_async(self.run_Blastp_Find, args=(bin_Folder, fasta_File, remote, species_Name, str(threads_per_loop)))
				elif stage == "compile_Database":
					pool.apply_async(self.compile_Database, args=(bin_Folder,))
				else:
					print("This Stage Does Not Exist")
					exit()
			else:
				print("Not Staging", stage, "for the Folder", bin_Folder)
		pool.close()
		pool.join()
	
	def convert_Gene_Clusters_to_Fasta(self, bin_Folder):
		print("Creating Fasta Files")
		# get File Info
		excel_Path = self.pangenome_Summary_Output + bin_Folder + "/"
		excel_File = excel_Path + self.project_Name + "_gene_clusters_summary_" + bin_Folder + ".xlsx"
		# Create Space for New Fasta Files
		place_Fasta_Files = excel_Path + "Fasta_Files/"
		self.execute_Command("mkdir " + place_Fasta_Files)
		# Load Excel Workbook
		WB = xl.load_workbook(excel_File) 
		WB_worksheets = WB.worksheets
		Main = WB_worksheets[0]
         		
		# NOTE: Everything in Excel is 1-indexed
		# Loop through Excel Spreedsheet Skipping Header
		for cell in Main['Q'][1:]:
			sheet_row = cell.row - 1 # 1-indexed to 0-indexed
			prot_Seq = cell.value
			if prot_Seq == None:
				continue
		 
			# Rename File
			bin_Name = Main["B"][sheet_row].value
			unique_ID = Main["A"][sheet_row].value
			strain_Name = Main["D"][sheet_row].value
			new_File = place_Fasta_Files + bin_Name + "_" + unique_ID + "_" + strain_Name + ".fasta"
		 	
		 	# Save as Fasta File
			records = []
			with open(new_File, "w+") as out_file:
				# Replace Unknown '-' as 'X'
				prot_Seq = prot_Seq.replace("-","X")
				record = SeqRecord(Seq(prot_Seq,IUPAC.protein), id=bin_Name)
				records.append(record)
				# Write the File
				SeqIO.write(records, new_File, "fasta")
				
	def run_Blastp_Find(self, bin_Folder, fasta_File, remote, species_Name, threads_per_loop):
		# Add the flag '-dbtype' to specify molecule type of input, values can be nucl or prot
		# Add the flag '-hash_index' to create index of sequence hash values
		# Add the flag '-parse_seqids' to parse bar delimited sequence identifiers (e.g., gi|129295) in FASTA input
		# get Output Path/Type
		excel_Path = self.pangenome_Summary_Output + bin_Folder + "/"
		base = os.path.splitext(fasta_File)[0]
		outfile = excel_Path + "blastp_Results/" + base + ".txt"
		outformat = "'6 qseqid sseqid pident mismatch gapopen evalue bitscore staxids sscinames scomnames sblastnames sskingdoms stitle'"
		# Specify Search Criteria
		max_Display = "5"; word_Size = "5"; evaluate = "1e-5";
		search_Database = "nr"   # refseq_protein, pdbaa, nt, nr
		short = " -task blastp-short "
		do_Remote = ""
		if species_Name:
			species_Name = " -entrez_query " + species_Name
		if remote:
			do_Remote = " -remote "
		# Blastp Search
		command = "blastp -parse_deflines -query " + excel_Path + "Fasta_Files/" + fasta_File + short + " -db " + search_Database + " " + do_Remote + " -word_size " + word_Size + " -evalue " + evaluate + " -outfmt " + outformat + " -out " + outfile + " -max_target_seqs " + max_Display + species_Name + " -num_threads " + threads_per_loop
		self.execute_Command(command)
				
	def compile_Database(self, bin_Folder):
		print("\nCompiling Datase")
		# get File Info
		blastp_Results = self.pangenome_Summary_Output + bin_Folder + "/blastp_Results/"
		database_Full = self.pangenome_Summary_Output + bin_Folder + "/genome_Dictionary_" + bin_Folder + "_Full.json"
		database_List_Names = self.pangenome_Summary_Output + bin_Folder + "/genome_Dictionary_" + bin_Folder + "_Protein_List.json"
		
		proteins = dict()
		dont_use = {database_Full,database_List_Names}
		# Loop Through the Blastp Output Files
		for blast_File in os.listdir(blastp_Results):
			if blast_File.endswith(".txt") and blast_File not in dont_use:
				# Open File from Blast
				with open(blastp_Results + blast_File, "r") as in_text:
					in_reader = csv.reader(in_text, delimiter = '\t')
					# Loop Through Blastp Results
					for row in in_reader:
						# Stop if No Data Present
						if len(row) == 0:
							continue
						# Only Trust Blast if Percent Identity is Above 95%
						percent_Identity = row[2]
						if float(percent_Identity) < 95:
							continue
						# Record What Blast Identified Protein as AND Keep Count of How Many Were Found
						protein_ID = row[12]
						if proteins.get(protein_ID, None) == None:
							proteins[protein_ID] = [1,blast_File]
							break
						proteins[protein_ID][0] += 1
						proteins[protein_ID].append(blast_File)
						# For now, We are ONLY Taking First Row (Best Find)
						break
			else:
				print("Not Using: ", blast_File)

		with open(database_Full, 'w') as outfile:
			json.dump(proteins, outfile, indent=4, sort_keys=True)
		
		with open(database_List_Names, 'w') as outfile:
			json.dump(list(proteins.keys()), outfile, indent=4, sort_keys=True)


# ----------------------------------------------------------------------------------------------------------------#

if __name__ == "__main__":
	"""
	Notes of Using this File:
		1. input_Fasta_File_Directory: 	This folder should have all the fasta files you want to make into a pangenome
								 		Files can end with ".fasta",  ".fna", ".fsa", ".mpfa", ".fa" (in output it will become '.fa')
								 		NOTHING in this folder will be edited (inside and out)
		2. output_File_Directory: 	All Files in 'fasta_Files/' are generated in data_Preparation class
									All Files in 'contigs_Database/' are generated in ANVIO_Storage class
									All Files in 'pan_Files/' are generated in Run_Pangenome class
									All Files in 'Helpers/' are generated in Additions_to_Pangenome class
									Hence, if you stop the program midway into a class, just delete the associated folder
	"""	
	# Gather User Inputs (From Command Line)
	get_Info = user_Inputs()
	input_Fasta_File_Directory, output_File_Directory, project_Name, pan_Title, threads = get_Info.gather_Inputs()
	# If No Inputs to Command Line: Use Inputs Below
	if len(sys.argv) == 1:
		input_Fasta_File_Directory = "./Input_File_Kalamiella_piersonii/"  	# Must end with '/'. Contains Fasta Files. NOTHING here will be edited
		output_File_Directory = "./Output_Files_Kalamiella_piersonii/"      	# Must end with '/'. Files could be edited (or not).
		project_Name = "Kalamiella_piersonii_Pangenome"		# Name of Project. NO SPACES (This is a Filename)
		pan_Title = "'Kalamiella piersonii Pangenome'"  		# Title of Graph. IF THERE ARE SPACES, you MUST ALSO wrap in single quotes (Ex: "'My Title'")
		threads = 5		                               		# specify the maximum number of threads you are willing to use on this program

		
	# Make Output Folders in output_File_Directory
	header = helper_Files(output_File_Directory)
	header.make_Output_Folders(output_File_Directory)

	# -----------------------------------------------PART ONE: Create Pangenome; No Human Needed-----------------------------------------------------------------#
	if False:
		# Data Preparation
		edit_FASTA_Files = data_Preparation(input_Fasta_File_Directory, output_File_Directory)  # Create Class Instance
		edit_FASTA_Files.remove_Extended_Nucleotides()  # Remove non-ATCGN Nucleotides (ANVIO Cannot Deal With Them)
		edit_FASTA_Files.fix_file_names()               # Edit Fasta File names to comply with ANVIO database

		# Create Database Storage with ANVIO
		Make_Storage = ANVIO_Storage(output_File_Directory, project_Name, threads)  # Create Class Instance.
		Make_Storage.storage_Step("format")    # Remove Bad Contigs and Format Fasta File to Work with Anvio's Software
		Make_Storage.storage_Step("convert")   # Convert FASTA '.fa' to '.db', YOU MUST MUST run 'anvi-setup-ncbi-cogs' once in the terminal
		Make_Storage.create_ANVIO_storage()	# Create a Storage File so ANVIO Can Know Where All the Contig Databases are Located
		
		# Make Pangenome
		make_Pan = Run_Pangenome(output_File_Directory, project_Name, threads) # Create Class Instance
		make_Pan.run_Pangenome()         # Create the Pangenome
		
		# Add Additional Material to Pangenome
		additions = Additions_to_Pangenome(output_File_Directory, project_Name, threads) # Create Class Instance
		additions.add_Genome_Clusters()  # Add Genome Clusters (Not Super Helpful)
		additions.add_beauty()			 # Add some pre-established beauty (You can still change it later)
		additions.add_ANI()			  	 # Add an ANI (You Can Remove Later). In Fact, You MUST Manually Check Display ANI on Pangenome
	
	# -----------------------------------------------PART TWO: Display Pangenome; Human Needed-----------------------------------------------------------------#
	if False:
		# Make Pangenome
		make_Pan = Run_Pangenome(output_File_Directory, project_Name, threads) # Create Class Instance
		make_Pan.display_Pangenome(pan_Title)  							       # Display the Pangenome
	
	# -----------------------------------------------PART Three: Identify Genes; No Human Needed-----------------------------------------------------------------#
	pangenome_Bin_Collection_Saved_As = "default"	# The Name you Saved the Pangenome Summary as (Under Bins Section)
	species_Name = ""            					# Leave as "" if you want to blast against all (or more than one species present). Example: "'Bacillus pumilus'" (dont forget BOTH quotes)
	remote = False
	if True:
		# Make Pangenome
		make_Pan = Run_Pangenome(output_File_Directory, project_Name, threads) # Create Class Instance
		#make_Pan.summarize_Pangenome(pangenome_Bin_Collection_Saved_As)   	   # Record the Bins you Saved AFTER OPENING and EDITING the pangenome
		
		# Identify Genes in Summarized Pangenome
		extract_Gene_Info = identify_Genes(output_File_Directory, project_Name, pangenome_Bin_Collection_Saved_As, threads) # Create Class Instance
		#extract_Gene_Info.compile_Excel()							# Compile Excel File with Pangenome Summary Information and Extract Bin Info
		extract_Gene_Info.divide_Work("create_Fasta_Files")			# Create Fasta Files of Genes Inside Bins 
		#extract_Gene_Info.divide_Work("find_Closest_Protein", remote, species_Name)	# Blastp the Fasta Files to Find the Identities of the Proteins in the Bins
		#extract_Gene_Info.divide_Work("compile_Database")			# Record the Protein IDs from Blasting the Genes in One Database for Each Bin
		
		
		
		
"""
		input_Fasta_File_Directory = "./Input_Files_Rhodotorula/"  # Must end with '/'. Contains Fasta Files. NOTHING here will be edited
		output_File_Directory = "./Output_Files_Rhodotorula/"      # Must end with '/'. Files could be edited (or not).
		project_Name = "Rhodotorula_mucilaginosa_Pangenome"	# Name of Project. NO SPACES (This is a Filename)
		pan_Title = "'Rhodotorula mucilaginosa Pangenome'"  # Title of Graph. IF THERE ARE SPACES, you MUST ALSO wrap in single quotes (Ex: "'My Title'")
		threads = 2		                               		# specify the maximum number of threads you are willing to use on this program
		
Trying to get to work:

    def add_Functions(self, category, annotation, func_file = None):
        #MUST have annotated the genomes on a previous step. In this function then import the layers you want
        
        # If you need to import a layer from func_file (else, you are suing the layers already given)
        # Confused about what a layer is here: open the pangenome image and look at the layer tab. This will be your catagory below
        if func_file != None:
            # First Import a Tab Delimited File with Functions
            command  = "anvi-import-misc-data " + func_file + " -p " + self.PanFile + " --target-data-table items"
            print("Executing the Following Command: ", command)
            ANVIO_functions = os.system(command)
        # --------------------------------------- #
        # YOU MUST change this to work with your files (or maybe you will get lucky)
        # NOTE: I have found that order matters in the command (specifically catagory before annotation)
        # Optonal: '--include-gc-identity-as-function'
        # Catagory: It is one of the names on the right most side of the pangenome (on the +x axis), You should have imported with func_file first
        # Annotation: it is one of the name in the center line of the pangenome (on the +y axis), like COG_FUNCTION
        functions_output = self.helper_Folder + "Klebsiella_pneumoniae_functions"
        command = "anvi-get-enriched-functions-per-pan-group -p " + self.PanFile + " -g " + self.genome_storage_database + " --include-gc-identity-as-function --category " + category + " --annotation-source " + annotation + " -o " + functions_output
        print("Executing the Following Command: ", command)
        ANVIO_functions = os.system(command)
"""
    
